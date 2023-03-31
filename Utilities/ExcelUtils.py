import openpyxl
from openpyxl.styles import PatternFill

def getRowCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_row)

def getColumnCount(file,sheetname):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return(sheet.max_column)

def readData(file,sheetname,rn,cn):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    return sheet.cell(rn,cn).value

def writeData(file,sheetname,rn,cn,data):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    sheet.cell(rn,cn).value=data
    workbook.save(file)

def greenFill(file,sheetname,rn,cn):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    greenFill= PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rn,cn).fill=greenFill
    workbook.save(file)

def redFill(file,sheetname,rn,cn):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook[sheetname]
    redFill= PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rn,cn).fill=redFill
    workbook.save(file)